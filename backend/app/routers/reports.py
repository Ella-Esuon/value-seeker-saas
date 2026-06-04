import io
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import date
from .. import models
from ..auth import get_tenant_id
from ..database import get_db

router = APIRouter(prefix="/api/reports", tags=["reports"])

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]


def _members_data(db, tenant_id):
    members = (
        db.query(models.Member)
        .filter(models.Member.tenant_id == tenant_id)
        .order_by(models.Member.name)
        .all()
    )
    rows = []
    for m in members:
        total = sum(c.amount for c in m.contributions)
        loans = [l for l in m.loans if l.status in ("active", "paid")]
        total_loans = sum(l.amount for l in loans)
        rows.append({
            "membership_id": m.membership_id, "name": m.name,
            "phone": m.phone or "", "email": m.email or "",
            "status": m.status, "date_joined": str(m.date_joined),
            "total_contributions": total, "total_loans": total_loans,
        })
    return rows


def _contributions_data(db, tenant_id, year):
    members = (
        db.query(models.Member)
        .filter(models.Member.tenant_id == tenant_id)
        .order_by(models.Member.name)
        .all()
    )
    rows = []
    for m in members:
        months_paid = {c.month: c.amount for c in m.contributions if c.year == year}
        year_total = sum(months_paid.values())
        row = {"name": m.name, "membership_id": m.membership_id}
        for mn in range(1, 13):
            row[MONTHS[mn-1]] = months_paid.get(mn, 0)
        row["total"] = year_total
        rows.append(row)
    return rows


def _loans_data(db, tenant_id):
    loans = (
        db.query(models.Loan)
        .filter(models.Loan.tenant_id == tenant_id)
        .order_by(models.Loan.date_applied.desc())
        .all()
    )
    rows = []
    for l in loans:
        repaid = sum(r.principal_paid for r in l.repayments)
        rows.append({
            "id": l.id, "member": l.member.name,
            "membership_id": l.member.membership_id,
            "amount": l.amount, "rate": l.interest_rate,
            "term": l.term_months, "installment": l.monthly_installment,
            "issued": str(l.date_issued) if l.date_issued else "",
            "due": str(l.due_date) if l.due_date else "",
            "status": l.status, "repaid": repaid,
            "balance": round(max(0, l.amount - repaid), 2),
        })
    return rows


def _delinquent_data(db, tenant_id):
    today = date.today()
    loans = db.query(models.Loan).filter(
        models.Loan.tenant_id == tenant_id,
        models.Loan.status == "active",
        models.Loan.due_date != None,
        models.Loan.due_date < today,
    ).all()
    rows = []
    for l in loans:
        repaid = sum(r.principal_paid for r in l.repayments)
        days = (today - l.due_date).days
        rows.append({
            "member": l.member.name, "membership_id": l.member.membership_id,
            "amount": l.amount, "balance": round(max(0, l.amount - repaid), 2),
            "due_date": str(l.due_date), "days_overdue": days,
        })
    return rows


# ── Excel ────────────────────────────────────────────────────────────────────

def _excel_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _style_header(ws, cols):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="1A3A5C")
    font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20


@router.get("/members/excel")
def members_excel(db: Session = Depends(get_db), tenant_id: int = Depends(get_tenant_id)):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"
    cols = ["Membership ID","Name","Phone","Email","Status","Date Joined","Total Contributions","Total Loans"]
    _style_header(ws, cols)
    for row in _members_data(db, tenant_id):
        ws.append([row["membership_id"],row["name"],row["phone"],row["email"],
                   row["status"],row["date_joined"],row["total_contributions"],row["total_loans"]])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    return _excel_response(wb, "members_report.xlsx")


@router.get("/contributions/excel")
def contributions_excel(
    year: int = Query(default=None),
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_tenant_id),
):
    from openpyxl import Workbook
    year = year or date.today().year
    wb = Workbook()
    ws = wb.active
    ws.title = f"Contributions {year}"
    cols = ["Membership ID","Name"] + MONTHS + ["Total"]
    _style_header(ws, cols)
    for row in _contributions_data(db, tenant_id, year):
        ws.append([row["membership_id"],row["name"]] + [row[m] for m in MONTHS] + [row["total"]])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 10
    return _excel_response(wb, f"contributions_{year}.xlsx")


@router.get("/loans/excel")
def loans_excel(db: Session = Depends(get_db), tenant_id: int = Depends(get_tenant_id)):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Portfolio"
    cols = ["ID","Member","Membership ID","Amount","Rate %","Term","Installment","Issued","Due","Status","Repaid","Balance"]
    _style_header(ws, cols)
    for row in _loans_data(db, tenant_id):
        ws.append([row["id"],row["member"],row["membership_id"],row["amount"],row["rate"],
                   row["term"],row["installment"],row["issued"],row["due"],
                   row["status"],row["repaid"],row["balance"]])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14
    return _excel_response(wb, "loan_portfolio.xlsx")


@router.get("/delinquent/excel")
def delinquent_excel(db: Session = Depends(get_db), tenant_id: int = Depends(get_tenant_id)):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Delinquent Loans"
    cols = ["Member","Membership ID","Loan Amount","Balance","Due Date","Days Overdue"]
    _style_header(ws, cols)
    for row in _delinquent_data(db, tenant_id):
        ws.append([row["member"],row["membership_id"],row["amount"],
                   row["balance"],row["due_date"],row["days_overdue"]])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16
    return _excel_response(wb, "delinquent_loans.xlsx")


# ── PDF ──────────────────────────────────────────────────────────────────────

def _pdf_response(buffer, filename):
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_pdf(title, col_widths, col_headers, rows_data):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph("VALUE SEEKER Financial Management", styles["Title"]))
    elements.append(Paragraph(title, styles["Heading2"]))
    elements.append(Paragraph(f"Generated: {date.today()}", styles["Normal"]))
    elements.append(Spacer(1, 0.4*cm))
    table_data = [col_headers] + rows_data
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1A3A5C")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#EEF2F7")]),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("PADDING",    (0,0), (-1,-1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    return buf


@router.get("/members/pdf")
def members_pdf(db: Session = Depends(get_db), tenant_id: int = Depends(get_tenant_id)):
    from reportlab.lib.units import cm
    rows = _members_data(db, tenant_id)
    col_widths = [3*cm,5*cm,3*cm,4*cm,2.5*cm,3*cm,3.5*cm,3*cm]
    headers = ["Membership ID","Name","Phone","Email","Status","Joined","Contributions","Loans"]
    data = [[r["membership_id"],r["name"],r["phone"],r["email"],
             r["status"],r["date_joined"],f"GH₵{r['total_contributions']:,.2f}",
             f"GH₵{r['total_loans']:,.2f}"] for r in rows]
    buf = _build_pdf("Member Report", col_widths, headers, data)
    return _pdf_response(buf, "members_report.pdf")


@router.get("/contributions/pdf")
def contributions_pdf(
    year: int = Query(default=None),
    db: Session = Depends(get_db),
    tenant_id: int = Depends(get_tenant_id),
):
    from reportlab.lib.units import cm
    year = year or date.today().year
    rows = _contributions_data(db, tenant_id, year)
    col_widths = [3*cm, 4*cm] + [1.8*cm]*12 + [3*cm]
    headers = ["ID","Name"] + MONTHS + ["Total"]
    data = [[r["membership_id"],r["name"]] + [f"{r[m]:,.0f}" if r[m] else "—" for m in MONTHS] +
            [f"GH₵{r['total']:,.2f}"] for r in rows]
    buf = _build_pdf(f"Annual Contribution Report — {year}", col_widths, headers, data)
    return _pdf_response(buf, f"contributions_{year}.pdf")


@router.get("/loans/pdf")
def loans_pdf(db: Session = Depends(get_db), tenant_id: int = Depends(get_tenant_id)):
    from reportlab.lib.units import cm
    rows = _loans_data(db, tenant_id)
    col_widths = [1.5*cm,4*cm,3*cm,3*cm,2*cm,2*cm,3*cm,3*cm,2.5*cm,3*cm]
    headers = ["ID","Member","Mbr ID","Amount","Rate","Term","Installment","Status","Repaid","Balance"]
    data = [[r["id"],r["member"],r["membership_id"],f"GH₵{r['amount']:,.2f}",
             f"{r['rate']}%",f"{r['term']}m",f"GH₵{r['installment']:,.2f}",
             r["status"],f"GH₵{r['repaid']:,.2f}",f"GH₵{r['balance']:,.2f}"] for r in rows]
    buf = _build_pdf("Loan Portfolio Report", col_widths, headers, data)
    return _pdf_response(buf, "loan_portfolio.pdf")


@router.get("/delinquent/pdf")
def delinquent_pdf(db: Session = Depends(get_db), tenant_id: int = Depends(get_tenant_id)):
    from reportlab.lib.units import cm
    rows = _delinquent_data(db, tenant_id)
    col_widths = [5*cm,3.5*cm,3.5*cm,3.5*cm,3*cm,3*cm]
    headers = ["Member","Membership ID","Loan Amount","Balance","Due Date","Days Overdue"]
    data = [[r["member"],r["membership_id"],f"GH₵{r['amount']:,.2f}",
             f"GH₵{r['balance']:,.2f}",r["due_date"],r["days_overdue"]] for r in rows]
    buf = _build_pdf("Delinquent Loan Report", col_widths, headers, data)
    return _pdf_response(buf, "delinquent_loans.pdf")
